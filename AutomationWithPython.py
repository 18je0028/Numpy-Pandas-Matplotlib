import openpyxl as xl
from openpyxl.chart import BarChart , Reference
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW, MIN_COLUMN, MIN_ROW

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

# To get the value at excel sheet at matrix (1,1)
# Method-1 (a1)-> ( meaning is -> column a and row 1)
cell = sheet['a1']
print(cell.value)

# Method-2
# To get the value at excel sheet at matrix (1,1)
cell = sheet.cell(1,1)
print(cell.value)

# To get total number of rows and columns in our excel sheet
print(sheet.max_row)
print(sheet.max_column)

# To extract data from a certain row
for row in range(1,sheet.max_row+1):
    cell = sheet.cell(row,3)
    print(cell.value)

# To create a new excel sheet and write the correct price of the item if 
# the price of all items is reduced by 90%

for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    correctedPrice = cell.value*0.9
    correctedCellPrice = sheet.cell(row,4)
    correctedCellPrice.value = correctedPrice



values = values = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('Transactions2.xlsx')

